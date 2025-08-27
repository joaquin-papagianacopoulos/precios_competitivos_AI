import openpyxl

def unmerge_excel(input_file: str, output_file: str):
    wb = openpyxl.load_workbook(input_file)
    for ws in wb.worksheets:
        merged_cells = list(ws.merged_cells)  # lista de celdas combinadas
        for merged in merged_cells:
            # Guardamos el valor de la celda superior izquierda
            min_row, min_col, max_row, max_col = merged.min_row, merged.min_col, merged.max_row, merged.max_col
            value = ws.cell(row=min_row, column=min_col).value

            # Descombinamos
            ws.unmerge_cells(str(merged))

            # Rellenamos todas las celdas con el mismo valor
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    ws.cell(row=row, column=col).value = value

    wb.save(output_file)
    print(f"✅ Archivo guardado sin celdas combinadas: {output_file}")

# ---------------------------
# Uso
# ---------------------------
input_file = "listas/labomba.xlsx"
output_file = "labomba_sin_combinadas.xlsx"
unmerge_excel(input_file, output_file)
